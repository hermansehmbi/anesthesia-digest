"""
Anesthesia Journal Digest — Self-Assessment Tracker (.xlsx)
===========================================================
A plain, macro-free workbook (works in Excel, Excel-on-web, and Google Sheets)
to help a physician self-report self-directed reading/reflection of the
anesthesia literature under RCPSC MOC Section 2 (Self-Learning).

Three sheets:
  1. "About"             — what it is + how to use it.
  2. "Activity Log"      — Date | Hours | Credits (self-reported) | Title |
                           Journal | APA Reference. User-fillable; add rows freely.
  3. "Summary & Report"  — print-ready totals with a "print for your MOC
                           submission" call-to-action banner.

Public API (unchanged for callers): MOC_FILE, get_or_create, log_articles,
log_cme, update_summary.
"""

import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

logger = logging.getLogger(__name__)
MOC_FILE = "self_assessment_tracker.xlsx"

# Sheet names
ABOUT_SHEET = "About"
LOG_SHEET = "Activity Log"
REPORT_SHEET = "Summary & Report"

# Activity Log column order (1-based): Date | Hours | Credits | Title | Journal | APA
LOG_HEADERS = ["Date", "Hours", "Credits (self-reported)", "Title", "Journal",
               "APA Reference"]
LOG_WIDTHS = [13, 9, 20, 52, 22, 70]
# Bounded ranges for the Summary formulas (plenty for years of weekly logging).
_LOG_MAXROW = 2000

# Palette
NAVY = "1A5276"
ACCENT = "C0392B"     # call-to-action banner (stands out; white text)
LIGHT = "EAF1F8"
RULE = "D5DBE1"

_thin = Side(style="thin", color=RULE)
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# ── APA reference helper ─────────────────────────────────────────────────────

def _apa_authors(raw: str) -> str:
    """Turn 'Weikai Chen; Yanjun Wang' into APA author list 'Chen, W., & Wang, Y.'"""
    names = [n.strip() for n in (raw or "").replace(" and ", ";").split(";") if n.strip()]
    out = []
    for n in names:
        parts = n.split()
        if len(parts) == 1:
            out.append(parts[0])
            continue
        surname = parts[-1]
        initials = " ".join(f"{p[0].upper()}." for p in parts[:-1] if p)
        out.append(f"{surname}, {initials}".strip())
    if not out:
        return ""
    if len(out) == 1:
        return out[0]
    if len(out) <= 20:
        return ", ".join(out[:-1]) + ", & " + out[-1]
    # APA 7: first 19, ellipsis, final author.
    return ", ".join(out[:19]) + ", … " + out[-1]


def apa_reference(art: dict) -> str:
    """Build a single-line APA-style reference from an article dict."""
    authors = _apa_authors(art.get("authors", ""))
    date_str = str(art.get("date_str") or art.get("date") or "")
    year = date_str[:4] if len(date_str) >= 4 and date_str[:4].isdigit() else "n.d."
    title = (art.get("title") or "").strip().rstrip(".")
    journal = art.get("journal") or art.get("journal_abbr") or ""
    doi = (art.get("doi") or "").strip()
    url = art.get("url") or ""
    link = f"https://doi.org/{doi}" if doi else url

    ref = ""
    if authors:
        ref += f"{authors} "
    ref += f"({year}). {title}. {journal}."
    if link:
        ref += f" {link}"
    return ref.strip()


# ── Workbook lifecycle ───────────────────────────────────────────────────────

def get_or_create(filepath: str = MOC_FILE) -> openpyxl.Workbook:
    if Path(filepath).exists():
        return openpyxl.load_workbook(filepath)
    return _create_new()


def log_articles(articles: list, filepath: str = MOC_FILE):
    """Append article rows to the Activity Log (Date | Hours | Credits | Title |
    Journal | APA Reference). Hours are left blank for the user; Credits is a
    self-reported formula (= Hours × 2, the Section 2 rate)."""
    wb = get_or_create(filepath)
    ws = wb[LOG_SHEET]

    for art in articles:
        r = ws.max_row + 1
        ws.cell(r, 1, value=art.get("date_str", datetime.now().strftime("%Y-%m-%d")))
        ws.cell(r, 2, value="")                       # Hours — user fills
        ws.cell(r, 3, value=f'=IF(B{r}="","",B{r}*2)')  # Credits (self-reported)
        ws.cell(r, 4, value=(art.get("title", "") or "")[:300])
        ws.cell(r, 5, value=art.get("journal", "") or art.get("journal_abbr", ""))
        ws.cell(r, 6, value=apa_reference(art))
        _style_log_row(ws, r)

    wb.save(filepath)
    logger.info(f"Logged {len(articles)} articles to Self-Assessment Tracker")


def log_cme(questions: list, filepath: str = MOC_FILE):
    """Append the weekly self-assessment as one Activity Log row, consistent with
    the article layout. Self-reportable under RCPSC MOC Section 2 (Self-Learning)."""
    wb = get_or_create(filepath)
    ws = wb[LOG_SHEET]

    today = datetime.now().strftime("%Y-%m-%d")
    journals = "; ".join(sorted({q.get("source_journal", "") for q in questions if q.get("source_journal")}))
    r = ws.max_row + 1
    ws.cell(r, 1, value=today)
    ws.cell(r, 2, value="")                       # Hours — user fills
    ws.cell(r, 3, value=f'=IF(B{r}="","",B{r}*2)')  # Credits (self-reported)
    ws.cell(r, 4, value=f"Weekly self-assessment — {len(questions)} questions")
    ws.cell(r, 5, value=journals or "Multiple journals")
    ws.cell(r, 6, value="Self-directed reading & reflection of the week's featured articles.")
    _style_log_row(ws, r)

    wb.save(filepath)
    logger.info("Logged weekly self-assessment to Self-Assessment Tracker")


def update_summary(filepath: str = MOC_FILE):
    """Totals on the report are live formulas, so this just refreshes the
    'last updated' stamp (and re-ensures the report exists)."""
    wb = get_or_create(filepath)
    if REPORT_SHEET not in wb.sheetnames:
        _build_report(wb)
    ws = wb[REPORT_SHEET]
    ws["B5"] = "Last updated: " + datetime.now().strftime("%Y-%m-%d %H:%M")
    wb.save(filepath)
    logger.info("Refreshed Self-Assessment Tracker summary")


# ── Row styling ──────────────────────────────────────────────────────────────

def _style_log_row(ws, r: int):
    for c in range(1, len(LOG_HEADERS) + 1):
        cell = ws.cell(r, c)
        cell.border = _BORDER
        cell.alignment = Alignment(
            vertical="top",
            horizontal="center" if c in (1, 2, 3) else "left",
            wrap_text=c in (4, 6),
        )


# ── Workbook construction ────────────────────────────────────────────────────

def _create_new() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    # Sheets are created in order: About (the default sheet, renamed), then
    # Activity Log, then Summary & Report.
    _build_about(wb, wb.active)
    _build_log(wb)
    _build_report(wb)
    return wb


def _build_about(wb, ws):
    ws.title = ABOUT_SHEET
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 100

    title = ws["B2"]
    title.value = "Self-Assessment Tracker"
    title.font = Font(bold=True, size=22, color=NAVY)
    ws["B3"].value = "Anesthesia Journal Digest"
    ws["B3"].font = Font(size=12, color="888888", italic=True)

    blocks = [
        ("What this is",
         "A simple log for your own self-directed study of the anesthesia literature. "
         "Use it to record the articles you read and reflect on each week, the time you "
         "spend, and the self-reported credits you choose to claim."),
        ("Purpose",
         "It helps you keep a tidy, printable record of self-learning that may be eligible "
         "for self-reporting under RCPSC MOC Section 2 (Self-Learning). Logging an activity "
         "here is a personal record only — you are responsible for confirming your own "
         "eligibility before submitting anything to the Royal College."),
        ("How to add entries",
         "Go to the “Activity Log” sheet. Each row is one activity. New article picks are "
         "added for you automatically; you can also add your own rows at the bottom. Fill in "
         "the Hours you spent — the Credits (self-reported) column fills in automatically at "
         "the Section 2 rate of 2 credits per hour (copy the formula down for rows you add)."),
        ("How the totals work",
         "The “Summary & Report” sheet adds up your hours, self-reported credits, and number "
         "of entries automatically from the Activity Log. There is nothing to recalculate by "
         "hand — edit the log and the totals update."),
        ("How to print your report",
         "Open the “Summary & Report” sheet and choose File ▸ Print. The print area is "
         "preset to a clean one-page report you can save as PDF and upload to your MOC "
         "society."),
    ]
    r = 5
    for heading, body in blocks:
        ws.cell(r, 2, value=heading).font = Font(bold=True, size=13, color=NAVY)
        r += 1
        bcell = ws.cell(r, 2, value=body)
        bcell.font = Font(size=11, color="333333")
        bcell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 60
        r += 2

    note = ws.cell(r, 2,
                   value="Credits are self-reported (not accredited). RCPSC Section 2 "
                         "(Self-Learning) is self-reported at the physician's discretion.")
    note.font = Font(italic=True, size=10, color="888888")
    note.alignment = Alignment(wrap_text=True, vertical="top")


def _build_log(wb):
    ws = wb.create_sheet(LOG_SHEET)
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor=NAVY)
    for i, (h, w) in enumerate(zip(LOG_HEADERS, LOG_WIDTHS), 1):
        cell = ws.cell(1, i, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"
    # Make the self-reported nature of credits unmistakable.
    ws["C1"].comment = Comment(
        "Self-reported credits under RCPSC MOC Section 2 (Self-Learning); not accredited. "
        "Auto-calculated as Hours × 2 (the Section 2 rate).", "Self-Assessment Tracker")
    return ws


def _build_report(wb):
    if REPORT_SHEET in wb.sheetnames:
        del wb[REPORT_SHEET]
    ws = wb.create_sheet(REPORT_SHEET)
    ws.sheet_view.showGridLines = False
    for col, w in (("A", 4), ("B", 34), ("C", 22), ("D", 22)):
        ws.column_dimensions[col].width = w

    # Call-to-action print banner (bold fill, white text) across B:D.
    ws.merge_cells("B2:D2")
    banner = ws["B2"]
    banner.value = "PRINT THIS PAGE FOR YOUR MOC SUBMISSION  →  File ▸ Print"
    banner.font = Font(bold=True, size=13, color="FFFFFF")
    banner.fill = PatternFill("solid", fgColor=ACCENT)
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 30

    # Report title
    ws.merge_cells("B4:D4")
    ws["B4"].value = "Self-Assessment Tracker — Summary & Report"
    ws["B4"].font = Font(bold=True, size=16, color=NAVY)
    ws.merge_cells("B5:D5")
    ws["B5"].value = "Last updated: " + datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["B5"].font = Font(size=10, color="888888", italic=True)
    ws["B5"].alignment = Alignment(horizontal="left")

    # Physician identification (fill-in)
    ws["B7"].value = "Physician name:"
    ws["B7"].font = Font(bold=True, color="333333")
    ws.merge_cells("C7:D7")
    ws["C7"].fill = PatternFill("solid", fgColor=LIGHT)
    ws["C7"].border = _BORDER
    ws["B8"].value = "Royal College ID:"
    ws["B8"].font = Font(bold=True, color="333333")
    ws.merge_cells("C8:D8")
    ws["C8"].fill = PatternFill("solid", fgColor=LIGHT)
    ws["C8"].border = _BORDER

    # Totals table (live formulas off the Activity Log)
    ref_h = f"'{LOG_SHEET}'!$B$2:$B${_LOG_MAXROW}"
    ref_c = f"'{LOG_SHEET}'!$C$2:$C${_LOG_MAXROW}"
    ref_d = f"'{LOG_SHEET}'!$A$2:$A${_LOG_MAXROW}"

    ws["B10"].value = "Activity summary"
    ws["B10"].font = Font(bold=True, size=13, color=NAVY)

    rows = [
        ("Total entries logged", f"=COUNTA({ref_d})"),
        ("Total hours", f"=SUM({ref_h})"),
        ("Total credits (self-reported)", f"=SUM({ref_c})"),
    ]
    r = 11
    for label, formula in rows:
        lc = ws.cell(r, 2, value=label)
        lc.font = Font(size=11, color="333333")
        lc.border = _BORDER
        lc.alignment = Alignment(horizontal="left")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        vc = ws.cell(r, 3, value=formula)
        vc.font = Font(bold=True, size=12, color=NAVY)
        vc.alignment = Alignment(horizontal="center")
        vc.border = _BORDER
        ws.cell(r, 4).border = _BORDER
        r += 1

    ws.cell(r + 1, 2,
            value="Self-reported under RCPSC MOC Section 2 (Self-Learning) — credits are "
                  "self-reported (not accredited). Confirm your own eligibility before "
                  "submitting to the Royal College.")
    ws.cell(r + 1, 2).font = Font(italic=True, size=10, color="888888")
    ws.cell(r + 1, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 1, start_column=2, end_row=r + 2, end_column=4)

    # Print-ready page setup + defined print area.
    ws.print_area = "A1:D18"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.5
    return ws
